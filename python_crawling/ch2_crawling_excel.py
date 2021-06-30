#크롤링한 데이터를 엑셀파일로 저장하기 - openpyxl 라이브러리 활용
import openpyxl

excel_file=openpyxl.Workbook() #Workbook()으로 엑셀 파일 생성
#엑셀 파일이 생성되면 default sheet가 생성되고, (즉 excel_file안에는 excel과 관련된 모든 정보/메서드들이 있는 하나의 객체가 담김.
# 따라서 그 객체 정보 중에서! sheet를 사용하려면 .active를 해주는 거임)
# 엑셀파일변수.active로 그 sheet를 다른 변수에 담아 활용
#- 해당 sheet 이름을 변경하려면 title변수값을 변경해주면 됨
''' 위의 예시:
  excel_sheet=excel_file.active #sheet를 쓰겠다
  excel_sheet.title='report'   #그 sheet의 title 변경
'''

excel_sheet=excel_file.active #객체 중 sheet 기능
#데이터 추가하기-가장 간단한 방법으로 sheet변수.append(리스트 형태의 하나의 열 데이터)로 추가해줌
excel_sheet.append(['data1','data2','data3']) #append는 행으로 저장됨. 즉 a1,b1,c1 순서로 data1,data2,data3이 추가됨 (a1,a2,a3가 아님!)
# 두 번 실행하면 이제 a2,b2,c2 칸에 데이터가 추가됨

#"파일" 저장-sheet가 아니라!file 자체를 저장
excel_file.save('tmp.xlsx')
#엑셀 파일 "닫기" - 프로그래밍에서 파일을 처리할 때는 반드시 파일을 닫아줘야 함!
excel_file.close()


#위의 과정을 하나의 함수로 만들기
def write_excel_template(filename,sheetname,listdata): #(저장하고 싶은 엑셀 파일이름,저장할 sheet name, 저장할 data(리스트의 요소들이 리스트인 형태))
    #listdata: [[list1],[list2],[list3]....] 이런 형태
    excel_file=openpyxl.Workbook()
    excel_sheet=excel_file.active

    excel_sheet.column_dimensions['A'].width=100

    if sheetname !='': #딱히 sheet이름 지정 안 한거면 그냥 default로 넘어가고, 지정한 게 있으면 그걸로 바꿔줌
        excel_sheet.title=sheetname

    for item in listdata:
        excel_sheet.append(item) #item도 하나의 리스트임
    
    excel_file.save(filename)
    excel_file.close()

    